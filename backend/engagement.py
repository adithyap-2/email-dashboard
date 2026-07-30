"""Load the read-only engagement sheet from the Excel workbook into the app DB.

Only the "Updated Engagement Sheet" tab is used (the current source of truth per
the business team). The sheet is INPUT ONLY — we never write rows back to it and
never invent engagement records. In production this same table would be
populated by an n8n workflow reading the SharePoint copy; loading from the local
.xlsx keeps local/dev identical without touching Microsoft Graph.

Quirks handled:
  * Organisation is only filled on the first row of each org group (visually
    merged in Excel) — we forward-fill it down.
  * "Next Followup date and message" packs a due date and a short message hook
    into one cell — we split them.
  * Dates are free text ("20th July", "Not specific") — see dates.parse_fuzzy_date.
"""

from __future__ import annotations

import os
import re
from datetime import date

import openpyxl

import db
from dates import parse_fuzzy_date

XLSX_PATH = os.environ.get(
    "ENGAGEMENT_XLSX",
    os.path.join(os.path.dirname(__file__), "..", "Engagement Sheet.xlsx"),
)
SHEET_NAME = os.environ.get("ENGAGEMENT_SHEET", "Updated Engagement Sheet")

# Header label -> our field name. Matched case-insensitively on a prefix so
# minor wording drift in the sheet ("Next Followup date and message hook") still
# lands in the right column.
_COLMAP = {
    "organisation": "organisation",
    "name": "contact_name",
    "topic of concern": "topic",
    "last day of contact": "last_contact_raw",
    "overall summary of discussion": "summary",
    "next steps": "next_steps",
    "next followup date and message": "next_followup_raw",
    "email": "email",
}


def _match_columns(header: list) -> dict[int, str]:
    """Map column index -> field name using the header row."""
    out: dict[int, str] = {}
    for idx, cell in enumerate(header):
        label = str(cell).strip().lower() if cell is not None else ""
        if not label:
            continue
        for prefix, field in _COLMAP.items():
            if label.startswith(prefix):
                out[idx] = field
                break
    return out


def _split_followup(raw: str) -> tuple[str, str]:
    """Split a 'Next Followup date and message' cell into (date_part, hook).

    The date typically leads ("29th July — send MoU reminder"); anything after
    the first separator is treated as the message hook. If there's no date, the
    whole string becomes the hook.
    """
    if not raw:
        return "", ""
    text = str(raw).strip()
    # Split on the first dash/colon/newline that separates date from message.
    m = re.split(r"\s*[—\-:\n]\s*", text, maxsplit=1)
    date_part = m[0].strip()
    hook = m[1].strip() if len(m) > 1 else ""
    return date_part, hook


def _iso(d: date | None) -> str | None:
    return d.isoformat() if d else None


def _clean(v) -> str:
    return str(v).strip() if v is not None else ""


def normalize_rows(raw_rows: list[dict], today: date | None = None) -> list[dict]:
    """Turn raw engagement rows (from the Excel sheet OR from n8n's SharePoint
    sync) into normalised DB records. Shared by both ingestion paths so the
    parsing rules (org forward-fill, follow-up split, fuzzy dates) stay identical.

    Each raw row is a dict with any of: organisation, contact_name, email, topic,
    last_contact_raw, next_followup_raw, summary, next_steps.
    """
    today = today or date.today()
    records: list[dict] = []
    current_org = ""

    for rec in raw_rows:
        org = _clean(rec.get("organisation"))
        if org:
            current_org = org  # start of a new org group (forward-fill down)
        org = org or current_org

        contact = _clean(rec.get("contact_name"))
        email = _clean(rec.get("email"))
        topic = _clean(rec.get("topic"))

        # Skip fully blank spacer rows.
        if not contact and not email and not topic:
            continue

        followup_date_raw, hook = _split_followup(rec.get("next_followup_raw"))
        last_raw = _clean(rec.get("last_contact_raw"))

        records.append({
            "organisation": org or None,
            "contact_name": contact or None,
            "email": email or None,
            "topic": topic or None,
            "last_contact_raw": last_raw or None,
            "last_contact_date": _iso(parse_fuzzy_date(last_raw, today)),
            "next_followup_raw": _clean(rec.get("next_followup_raw")) or None,
            "next_followup_date": _iso(parse_fuzzy_date(followup_date_raw, today)),
            "message_hook": hook or None,
            "summary": _clean(rec.get("summary")) or None,
            "next_steps": _clean(rec.get("next_steps")) or None,
        })
    return records


def replace_engagements(records: list[dict]) -> int:
    """Full-refresh the engagements table (it is a read-only mirror — we never
    append, we replace). Returns the row count written."""
    with db.get_conn() as conn:
        conn.execute("DELETE FROM engagements")
        conn.executemany(
            """INSERT INTO engagements
               (organisation, contact_name, email, topic, last_contact_raw,
                last_contact_date, next_followup_raw, next_followup_date,
                message_hook, summary, next_steps)
               VALUES
               (:organisation, :contact_name, :email, :topic, :last_contact_raw,
                :last_contact_date, :next_followup_raw, :next_followup_date,
                :message_hook, :summary, :next_steps)""",
            records,
        )
    return len(records)


def load_engagements(today: date | None = None) -> int:
    """Read the Excel sheet and replace the engagements table. Returns the row
    count loaded. Missing file / sheet is non-fatal (returns 0) so the dashboard
    still boots on machines without the workbook."""
    today = today or date.today()
    path = os.path.abspath(XLSX_PATH)
    if not os.path.exists(path):
        return 0

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        return 0
    ws = wb[SHEET_NAME]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return 0

    colmap = _match_columns(rows[0])
    raw_dicts = [
        {field: raw_row[idx] for idx, field in colmap.items() if idx < len(raw_row)}
        for raw_row in rows[1:]
    ]
    records = normalize_rows(raw_dicts, today)
    replace_engagements(records)
    return len(records)